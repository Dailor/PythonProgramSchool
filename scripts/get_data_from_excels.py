import os
import json

from itertools import chain

from openpyxl import Workbook, load_workbook

from models import Group, Teacher, Pupil

path_to_excels = "classes"
result_filename = 'groups.json'
subject_need = 'Информатика 1 ч'
forbidden_words = 'СО', 'чтв'

marks_row = 8
marks_col = 3

teachers_row = 3
teachers_col = 1

path_to_json = ('to_excel', 'result_work.json')
path_to_excel = ('ready_in_excel',)

path_to_json = os.path.join(*path_to_json)
path_to_excel = os.path.join(*path_to_excel)


def get_all_files_names():
    return os.listdir(path_to_excels)


def check_col_on_forbidden_words(sheet, cell):
    if sheet.cell(row=marks_row - 3, column=cell.column).value:
        return any(word in sheet.cell(row=marks_row - 3, column=cell.column).value for word in forbidden_words)
    return True


def save_data(data):
    with open(result_filename, 'w') as f:
        f.write(json.dumps(data))


def get_teachers(sheet):
    return [Teacher(*teacher.split()[:2]) for teacher in
            sheet.cell(row=teachers_row, column=teachers_col).value.strip().split(':')[-1].split(', ')]


def get_teacher_name(sheet, cell):
    return sheet.cell(row=marks_row - 2, column=cell.column).value.strip().split()[-1]


def get_group_by_teacher_first_letters(g1, g2, short_name):
    name, surname = short_name.lower().split('.')
    t1_n, t1_s = g1.teacher.name[0].lower(), g1.teacher.surname[0].lower()

    if (t1_n == name and t1_s == surname) or (t1_s == name and t1_n == surname):
        return g1

    return g2


def add_data_to_sheet(sheet, data):
    for row_data in chain([data[0].keys()], map(lambda x: x.values(), data)):
        sheet.append(list(row_data))


def results_to_excel(data):
    for filename, data_in in data.items():
        filename, ext = os.path.splitext(filename)

        wb = Workbook()

        if filename == 'groups':
            for group_index in range(len(data_in)):
                group = data_in[group_index]
                add_data_to_sheet(wb.create_sheet(title=group['name'], index=group_index), group['pupils'])
        else:
            add_data_to_sheet(wb.create_sheet(title=filename, index=0), data_in)

        path_to_result_file = os.path.join(path_to_excel, filename + '.xlsx')
        wb.save(path_to_result_file)


def excel_to_json():
    all_excel_files = get_all_files_names()
    groups = list()
    undefined_pupils = list()

    for excel_file in all_excel_files:
        path_to_excel_file = os.path.join(path_to_excels, excel_file)
        wb = load_workbook(filename=path_to_excel_file)
        sheet = wb[subject_need]

        class_name = excel_file.split('.')[0].rsplit('_')[-1]

        group_1 = Group(class_name, 1)
        group_2 = Group(class_name, 2)

        teacher1, teacher2 = get_teachers(sheet)

        group_1.teacher = teacher1
        group_2.teacher = teacher2

        for row in sheet.iter_rows(min_row=marks_row, min_col=marks_col - 1):
            row = iter(row)
            surname, name = next(row).value.strip().split()
            pupil = Pupil(surname=surname, name=name)

            for cell in row:
                if cell.value and isinstance(cell.value, int) and check_col_on_forbidden_words(sheet, cell):
                    teacher_name = get_teacher_name(sheet, cell)
                    get_group_by_teacher_first_letters(group_1, group_2, teacher_name).append(pupil)
                    break

            if not pupil.group_set:
                pupil.class_name = class_name
                group_1.append(pupil)
                undefined_pupils.append(pupil)

        groups.extend([group_1, group_2])

    result_dict = Group.serialize_groups(groups)
    result_dict['teachers'] = [teacher.serialize() for teacher in Teacher.get_all_teachers()]
    result_dict['undefined_pupils'] = [pupil.serialize() for pupil in undefined_pupils]

    save_data(result_dict)

    return result_dict


def main():
    data = excel_to_json()
    results_to_excel(data)


if __name__ == '__main__':
    main()
    print("END")
